#!/usr/bin/env python3
"""Mutation tests for the published procurement workbook boundary."""

from __future__ import annotations

import copy
import hashlib
from pathlib import Path
import shutil
from tempfile import TemporaryDirectory
import unittest
from unittest.mock import patch

from openpyxl import load_workbook

from public_privacy_guard import private_identifier_fingerprint
from validate_paid_data_procurement import (
    CURRENT_DASHBOARD_VERSION,
    OUTPUT_XLSX,
    EXPECTED_XLSX_SHA256,
    read_json,
    scan_public_workbook_text,
    SOURCE_PATH,
    validate_workbook,
    WORKBOOK_SNAPSHOT_AS_OF,
    WORKBOOK_SNAPSHOT_VERSION,
)


class PaidDataWorkbookPrivacyTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.source = read_json(SOURCE_PATH)

    def validate_mutation(
        self,
        coordinate: str,
        value: str,
        fingerprints: frozenset[tuple[int, str]] | None = None,
    ) -> list[str]:
        with TemporaryDirectory() as directory:
            path = Path(directory) / "mutated.xlsx"
            shutil.copyfile(OUTPUT_XLSX, path)
            workbook = load_workbook(path, read_only=False, data_only=False)
            workbook["Response Scorecard"][coordinate] = value
            workbook.save(path)
            workbook.close()
            reviewed_hash = hashlib.sha256(path.read_bytes()).hexdigest()
            errors: list[str] = []
            contexts = [
                patch("validate_paid_data_procurement.OUTPUT_XLSX", path),
                patch(
                    "validate_paid_data_procurement.EXPECTED_XLSX_SHA256",
                    reviewed_hash,
                ),
            ]
            if fingerprints is not None:
                contexts.append(
                    patch(
                        "public_privacy_guard.PRIVATE_IDENTIFIER_FINGERPRINTS",
                        fingerprints,
                    )
                )
            with contexts[0], contexts[1]:
                if len(contexts) == 3:
                    with contexts[2]:
                        validate_workbook(copy.deepcopy(self.source), errors)
                else:
                    validate_workbook(copy.deepcopy(self.source), errors)
            return errors

    def test_rejects_email_in_reviewer_note(self) -> None:
        errors = self.validate_mutation("V14", "analyst@example.test")
        self.assertTrue(any("email address" in error for error in errors), errors)

    def test_current_public_outreach_states_are_fail_closed(self) -> None:
        outreach = {
            item["itemId"]: item
            for item in self.source["outreach"]
        }
        self.assertEqual(
            outreach["ecig-global-market-database"]["state"],
            "followup_sent_response_pending",
        )
        self.assertIn(
            "first follow-up on 2026-07-28",
            outreach["ecig-global-market-database"]["noteEn"],
        )
        self.assertEqual(
            outreach["euromonitor-passport-nicotine"]["state"],
            "expanded_schema_and_package_quotes_review_pending",
        )
        self.assertIn(
            "expanded numerical Germany sample",
            outreach["euromonitor-passport-nicotine"]["noteEn"],
        )
        self.assertIn(
            "indicative annual package quotes",
            outreach["euromonitor-passport-nicotine"]["noteEn"],
        )
        self.assertIn(
            "78-market e-vapour value-coverage list",
            outreach["euromonitor-passport-nicotine"]["noteEn"],
        )
        self.assertIn(
            "later eight-tab category-schema workbook",
            outreach["euromonitor-passport-nicotine"]["noteEn"],
        )
        self.assertIn(
            "lists 95 geographies",
            outreach["euromonitor-passport-nicotine"]["noteEn"],
        )
        self.assertIn(
            "country-year value cells are blank",
            outreach["euromonitor-passport-nicotine"]["noteEn"],
        )
        self.assertIn(
            "2026-07-29 call was completed",
            outreach["euromonitor-passport-nicotine"]["noteEn"],
        )
        self.assertIn(
            "conditional paid arrangement",
            outreach["euromonitor-passport-nicotine"]["noteEn"],
        )
        self.assertIn(
            "No extract, order, invoice, fee, subscription or commitment is authorised or accepted",
            outreach["euromonitor-passport-nicotine"]["noteEn"],
        )
        self.assertIn(
            "lender/buyer NDA data-room rights",
            outreach["euromonitor-passport-nicotine"]["noteEn"],
        )
        self.assertIn(
            "complete all-in terms remain open",
            outreach["euromonitor-passport-nicotine"]["noteEn"],
        )
        self.assertNotIn("pending and has not been sent", outreach["euromonitor-passport-nicotine"]["noteEn"])
        self.assertNotIn("meeting", outreach["euromonitor-passport-nicotine"]["noteEn"].lower())
        self.assertNotIn("account history", outreach["euromonitor-passport-nicotine"]["noteEn"].lower())
        self.assertNotIn("single-user", outreach["euromonitor-passport-nicotine"]["noteEn"].lower())
        self.assertNotIn("consultant", outreach["euromonitor-passport-nicotine"]["noteEn"].lower())
        self.assertNotIn("CEO", outreach["euromonitor-passport-nicotine"]["noteEn"])
        self.assertIn(
            "0/6 passes",
            outreach["euromonitor-passport-nicotine"]["noteEn"],
        )
        self.assertTrue(outreach["euromonitor-passport-nicotine"]["noteEn"].endswith("NOT SCORED."))
        self.assertNotRegex(
            outreach["euromonitor-passport-nicotine"]["noteEn"],
            r"(?:EUR|USD|GBP)\s*[0-9]",
        )
        self.assertIn("remains non-testable", outreach["euromonitor-passport-nicotine"]["noteEn"])
        self.assertEqual(
            outreach["circana-us-tobacco-pilot"]["state"],
            "administrative_qualification_received",
        )
        self.assertIn(
            "commercial qualification response was received",
            outreach["circana-us-tobacco-pilot"]["noteEn"].lower(),
        )
        self.assertIn(
            "same-thread clarification was sent on 2026-07-28",
            outreach["circana-us-tobacco-pilot"]["noteEn"],
        )
        self.assertIn(
            "Sample data, methodology and a non-binding quote remain pending",
            outreach["circana-us-tobacco-pilot"]["noteEn"],
        )
        self.assertIn(
            "NOT SCORED; no purchase, fee or commitment is authorised",
            outreach["circana-us-tobacco-pilot"]["noteEn"],
        )
        self.assertNotRegex(
            outreach["circana-us-tobacco-pilot"]["noteEn"],
            r"(?:EUR|USD|GBP)\s*[0-9]",
        )

    def test_dashboard_and_reviewed_daily_workbook_share_v35_snapshot(self) -> None:
        self.assertEqual(self.source["version"], CURRENT_DASHBOARD_VERSION)
        self.assertEqual(self.source["asOf"], WORKBOOK_SNAPSHOT_AS_OF)
        self.assertEqual(CURRENT_DASHBOARD_VERSION, WORKBOOK_SNAPSHOT_VERSION)
        self.assertEqual(
            hashlib.sha256(OUTPUT_XLSX.read_bytes()).hexdigest(),
            EXPECTED_XLSX_SHA256,
        )
        workbook = load_workbook(OUTPUT_XLSX, read_only=True, data_only=False)
        try:
            release_boundary = workbook["Decision"]["A3"].value
        finally:
            workbook.close()
        self.assertIn(f"Version {WORKBOOK_SNAPSHOT_VERSION}", release_boundary)
        self.assertIn(f"Verified {WORKBOOK_SNAPSHOT_AS_OF}", release_boundary)
        self.assertIn(f"Version {CURRENT_DASHBOARD_VERSION}", release_boundary)

    def test_rejects_private_path_in_reviewer_note(self) -> None:
        errors = self.validate_mutation("V14", "/Users/example/private/reply.eml")
        self.assertTrue(any("local or private path" in error for error in errors), errors)

    def test_rejects_uuid_and_message_metadata(self) -> None:
        value = "Message-ID: 123e4567-e89b-42d3-a456-426614174000"
        errors = self.validate_mutation("V14", value)
        self.assertTrue(any("UUID-like" in error for error in errors), errors)
        self.assertTrue(
            any("message, form or thread metadata" in error for error in errors),
            errors,
        )
        self.assertTrue(any("correspondence header" in error for error in errors), errors)

    def test_rejects_fingerprinted_private_counterparty(self) -> None:
        marker = "Example Confidential Workbook Counterparty"
        fingerprints = frozenset({private_identifier_fingerprint(marker)})
        errors = self.validate_mutation("V14", marker, fingerprints)
        self.assertTrue(
            any("private identifier fingerprint" in error for error in errors),
            errors,
        )

    def test_rejects_false_public_outreach_state(self) -> None:
        errors = self.validate_mutation("D14", "RESPONSE RECEIVED · READY FOR PURCHASE")
        self.assertTrue(any("public state" in error for error in errors), errors)

    def test_rejects_overstated_euromonitor_state(self) -> None:
        errors = self.validate_mutation("D15", "SAMPLE AND QUOTE RECEIVED")
        self.assertTrue(any("public state" in error for error in errors), errors)

    def test_rejects_overstated_euromonitor_boundary_note(self) -> None:
        errors = self.validate_mutation("X15", "Commercial evidence received")
        self.assertTrue(any("boundary differs" in error for error in errors), errors)

    def test_rejects_unsupported_public_boundary_note(self) -> None:
        errors = self.validate_mutation("X14", "Unsupported commercial claim")
        self.assertTrue(any("boundary differs" in error for error in errors), errors)

    def test_rejects_overstated_circana_state(self) -> None:
        errors = self.validate_mutation("D17", "SAMPLE AND QUOTE RECEIVED")
        self.assertTrue(any("public state" in error for error in errors), errors)

    def test_rejects_unsupported_circana_boundary_note(self) -> None:
        errors = self.validate_mutation("X17", "Commercial evidence received")
        self.assertTrue(any("boundary differs" in error for error in errors), errors)

    def test_relationship_text_uses_same_privacy_guard(self) -> None:
        errors: list[str] = []
        scan_public_workbook_text(
            "OOXML relationship fixture",
            "file:///Users/example/private/reply.eml",
            errors,
        )
        self.assertTrue(any("local or private path" in error for error in errors), errors)


if __name__ == "__main__":
    unittest.main()
