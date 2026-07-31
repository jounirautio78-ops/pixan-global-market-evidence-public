#!/usr/bin/env python3
"""Fail-closed validation for the public paid-data procurement shortlist."""

from __future__ import annotations

import csv
from datetime import date
import hashlib
import io
import json
from pathlib import Path
import re
import sys
from typing import Any
from urllib.parse import urlparse
import zipfile
from xml.etree import ElementTree

from openpyxl import load_workbook

from public_privacy_guard import contains_private_identifier
from build_paid_data_procurement import (
    CSV_FIELDS,
    OUTPUT_CSV,
    OUTPUT_JSON,
    OUTPUT_XLSX,
    SOURCE_PATH,
    normalised,
    render_csv,
    render_json,
    score_item,
)


EXPECTED_TOP_LEVEL = {
    "schemaVersion",
    "programmeId",
    "asOf",
    "version",
    "status",
    "publicBoundaryEn",
    "publicBoundaryFi",
    "rankingBoundaryEn",
    "rankingBoundaryFi",
    "currencyNoteEn",
    "currencyNoteFi",
    "outreach",
    "weights",
    "packageOptions",
    "items",
    "goCriteria",
    "stopCriteria",
    "avoidAsPrimary",
}
WEIGHT_IDS = {
    "annualSalesFit",
    "vapeSpecificity",
    "countryCoverage",
    "methodTransparency",
    "transactionLicenceFit",
    "costCertainty",
}
PRICE_TYPES = {"public_list_price", "vendor_quote"}
OUTREACH_KEYS = {"itemId", "state", "recordedOn", "noteEn", "noteFi"}
EXPECTED_OUTREACH = {
    "ecig-global-market-database": "followup_sent_response_pending",
    "euromonitor-passport-nicotine": "expanded_schema_and_package_quotes_review_pending",
    "niq-rms-pilot": "blocked_not_submitted",
    "circana-us-tobacco-pilot": "administrative_qualification_received",
}
EXPECTED_OUTREACH_NOTES = {
    "ecig-global-market-database": (
        "A free review-sample and non-binding quote request was sent on 2026-07-23, followed by the first follow-up on 2026-07-28. No response, sample, quote, data, method, coverage, licence, price or commitment is recorded. NOT SCORED; no purchase is authorised.",
        "Maksuton tarkistusnäyte ja ei-sitova tarjous pyydettiin 23.7.2026, ja ensimmäinen seuranta lähetettiin 28.7.2026. Vastausta, näytettä, tarjousta, dataa, menetelmää, kattavuutta, lisenssiä, hintaa tai sitoumusta ei ole kirjattu. EI PISTEYTETTY; ostoa ei ole valtuutettu.",
    ),
    "euromonitor-passport-nicotine": (
        "An expanded numerical Germany sample, a 78-market e-vapour value-coverage list, generic methodology, standard licence terms, indicative annual package quotes and a later eight-tab category-schema workbook were received by 2026-07-27. A 2026-07-29 call was completed. The vendor then offered a limited Germany extract under a conditional paid arrangement: the extract fee would be waived only if a wider country package were purchased within the stated window; otherwise the extract could be invoiced. No extract, order, invoice, fee, subscription or commitment is authorised or accepted. The private 2023–2024 Germany comparison remains non-testable because the product scope, retail-versus-tax stage and record-status bridge is unresolved. The later workbook lists 95 geographies but its country-year value cells are blank, and the 95-geography structure remains unreconciled to the 78-country quote. A populated current Germany 2022–2025 test, exact country-product-field-year coverage, lender/buyer NDA data-room rights and complete all-in terms remain open. All six gates are evaluated, but 0/6 passes. NOT SCORED.",
        "Laajennettu numeerinen Saksa-näyte, 78 sähkötupakkamarkkinan arvotietojen peittolista, yleinen menetelmäkuvaus, vakiolisenssiehdot, suuntaa-antavat vuosipakettitarjoukset ja myöhempi kahdeksan välilehden kategoriaskeematyökirja saatiin 27.7.2026 mennessä. Puhelu pidettiin 29.7.2026. Toimittaja tarjosi sen jälkeen rajattua Saksa-otetta ehdollisella maksullisella järjestelyllä: otteen maksu poistuisi vain, jos laajempi maapaketti ostettaisiin ilmoitetun määräajan kuluessa; muutoin ote voitaisiin laskuttaa. Otetta, tilausta, laskua, maksua tai sitoumusta ei ole valtuutettu tai hyväksytty. Yksityinen vuosien 2023–2024 Saksa-vertailu ei ole vielä testattavissa, koska tuoterajaus, vähittäis- ja verovaiheen ero sekä tietueiden tilasilta ovat ratkaisematta. Myöhempi työkirja luettelee 95 maantiedettä, mutta sen maa–vuosi-arvosolut ovat tyhjiä eikä 95 maantieteen rakennetta ole täsmäytetty 78 maan tarjoukseen. Täytetty ajantasainen Saksan 2022–2025-testi, täsmällinen maa–tuote–kenttä–vuosi-peitto, lainanantaja-/ostaja-NDA-datahuoneoikeudet ja täydelliset kaikki kustannukset kattavat ehdot ovat avoinna. Kaikki kuusi porttia on arvioitu, mutta 0/6 läpäisee. EI PISTEYTETTY.",
    ),
    "niq-rms-pilot": (
        "Not submitted: the available form requires acceptance of Terms of Use. No terms were accepted.",
        "Ei lähetetty: käytettävissä oleva lomake edellyttää käyttöehtojen hyväksymistä. Ehtoja ei hyväksytty.",
    ),
    "circana-us-tobacco-pilot": (
        "A commercial qualification response was received and a same-thread clarification was sent on 2026-07-28. Sample data, methodology and a non-binding quote remain pending. NOT SCORED; no purchase, fee or commitment is authorised.",
        "Kaupallinen kvalifiointivastaus saatiin ja samaan viestiketjuun lähetettiin täsmennys 28.7.2026. Näyteaineisto, menetelmä ja ei-sitova tarjous ovat edelleen saamatta. EI PISTEYTETTY; ostoa, maksua tai sitoumusta ei ole valtuutettu.",
    ),
}
FORBIDDEN_PUBLIC_TEXT = (
    "/users/",
    "file://",
    "submissionguid",
    "submission guid",
)
EMAIL_PATTERN = re.compile(r"\b[a-z0-9._%+-]+@[a-z0-9.-]+\.[a-z]{2,}\b", re.IGNORECASE)
UUID_PATTERN = re.compile(
    r"\b[0-9a-f]{8}-[0-9a-f]{4}-[1-5][0-9a-f]{3}-[89ab][0-9a-f]{3}-[0-9a-f]{12}\b",
    re.IGNORECASE,
)
PRIVATE_PATH_PATTERN = re.compile(
    r"(?i)(?:file:(?://|\\/\\/)|/(?:Users|home|private|tmp|var|etc)/|"
    r"[A-Z]:\\\\(?:Users|home|private|tmp|var|etc)\\\\)"
)
CORRESPONDENCE_HEADER_PATTERN = re.compile(
    r"(?im)^\s*(?:from|to|cc|bcc|subject|sent|date|message[-_ ]?id|thread[-_ ]?id)\s*:"
)
PRIVATE_METADATA_MARKERS = (
    "submissionguid",
    "formguid",
    "messageid",
    "threadid",
)
CURRENT_DASHBOARD_VERSION = "2026.07.31-37"
WORKBOOK_SNAPSHOT_VERSION = "2026.07.31-37"
WORKBOOK_SNAPSHOT_AS_OF = "2026-07-31"
EXPECTED_XLSX_SHA256 = "5b243335fb8ed64e7af5bcc935d2d7072af7ca55de682adf513aade7c0f7c6c9"
EXPECTED_RESPONSE_ROWS = (
    (
        "ecig-global-market-database",
        "ECigIntelligence",
        "Global Market Database",
        "REQUEST + FOLLOW-UP SENT · RESPONSE PENDING · NOT SCORED\nFI: PYYNTÖ + SEURANTA LÄHETETTY · VASTAUS ODOTTAA · EI PISTEYTETTY",
        None,
        "='Sources'!C6",
        "Status only. Request sent 2026-07-23 and first follow-up sent 2026-07-28. No response, sample, quote, data, method, coverage, licence, price or commitment is recorded. NOT SCORED; no purchase is authorised.",
    ),
    (
        "euromonitor-passport-nicotine",
        "Euromonitor International",
        "Passport Nicotine / e-vapour country series",
        "CALL COMPLETED + CONDITIONAL PAID GERMANY EXTRACT OFFERED · NOT ACCEPTED · 0/6 GATES PASS · NOT SCORED\nFI: PUHELU PIDETTY + EHDOLLINEN MAKSULLINEN SAKSA-OTE TARJOTTU · EI HYVÄKSYTTY · 0/6 PORTTIA LÄPÄISTY · EI PISTEYTETTY",
        None,
        "='Sources'!C9",
        "Status only. A conditional paid Germany extract was offered after the 2026-07-29 call. It has not been accepted or activated; no order, invoice, fee, subscription or commitment is authorised. Product scope, transaction-stage reconciliation, current Germany coverage, rights and all-in terms remain open. NOT SCORED; 0/6 gates pass.",
    ),
    (
        "niq-rms-pilot",
        "NielsenIQ",
        "Retail Measurement Services pilot",
        "NOT SUBMITTED · TERMS GATE\nFI: EI LÄHETETTY · EHTOPORTTI",
        None,
        "='Sources'!C12",
        "Status only. No response content or unlicensed data.",
    ),
    (
        "circana-us-tobacco-pilot",
        "Circana",
        "US Tobacco POS pilot",
        "FOLLOW-UP SENT 2026-07-28 · SAMPLE + QUOTE PENDING · NOT SCORED\nFI: SEURANTA LÄHETETTY 28.7.2026 · NÄYTE + TARJOUS ODOTTAVAT · EI PISTEYTETTY",
        None,
        "='Sources'!C13",
        "Status only. A direct follow-up requested a United States retail sample, channel and method details, a non-binding minimum configuration and transaction-use rights. A substantive sample and quote remain pending. NOT SCORED; no purchase or other commitment is authorised.",
    ),
)


def read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def valid_iso_date(value: Any) -> bool:
    if not isinstance(value, str) or not re.fullmatch(r"\d{4}-\d{2}-\d{2}", value):
        return False
    try:
        date.fromisoformat(value)
    except ValueError:
        return False
    return True


def valid_public_url(value: Any) -> bool:
    if not isinstance(value, str):
        return False
    parsed = urlparse(value)
    return (
        parsed.scheme == "https"
        and bool(parsed.hostname)
        and not parsed.username
        and not parsed.password
        and not any(
            key.casefold() in {"access_token", "api_key", "key", "password", "secret", "token"}
            for key, _ in __import__("urllib.parse").parse.parse_qsl(parsed.query)
        )
    )


def text_values(value: Any):
    if isinstance(value, str):
        yield value
    elif isinstance(value, dict):
        for item in value.values():
            yield from text_values(item)
    elif isinstance(value, list):
        for item in value:
            yield from text_values(item)


def scan_public_workbook_text(label: str, value: str, errors: list[str]) -> None:
    folded = re.sub(r"[^a-z0-9]+", "", value.casefold())
    if EMAIL_PATTERN.search(value):
        errors.append(f"{label}: paid-data XLSX contains an email address")
    if UUID_PATTERN.search(value):
        errors.append(f"{label}: paid-data XLSX contains a UUID-like private reference")
    if PRIVATE_PATH_PATTERN.search(value):
        errors.append(f"{label}: paid-data XLSX contains a local or private path")
    if any(marker in folded for marker in PRIVATE_METADATA_MARKERS):
        errors.append(f"{label}: paid-data XLSX contains private message, form or thread metadata")
    if CORRESPONDENCE_HEADER_PATTERN.search(value):
        errors.append(f"{label}: paid-data XLSX contains correspondence header text")
    if contains_private_identifier(value):
        errors.append(f"{label}: paid-data XLSX contains a private identifier fingerprint")


def validate_source(source: Any, errors: list[str]) -> None:
    if not isinstance(source, dict):
        errors.append("source must contain an object")
        return
    if set(source) != EXPECTED_TOP_LEVEL or source.get("schemaVersion") != 1:
        errors.append("source has an unsupported top-level schema")
        return
    if source.get("status") != "decision_support_only_no_purchase_authorised":
        errors.append("source must state that no purchase is authorised")
    if not valid_iso_date(source.get("asOf")) or source.get("version") != CURRENT_DASHBOARD_VERSION:
        errors.append("source date or version is invalid")
    if source.get("asOf") != WORKBOOK_SNAPSHOT_AS_OF:
        errors.append("dashboard and retained workbook must remain within one daily snapshot date")

    weights = source.get("weights")
    if not isinstance(weights, list) or len(weights) != 6:
        errors.append("exactly six transparent weights are required")
        return
    if {item.get("id") for item in weights if isinstance(item, dict)} != WEIGHT_IDS:
        errors.append("weight IDs differ from the reviewed score model")
    try:
        if abs(sum(float(item["weight"]) for item in weights) - 1.0) > 1e-9:
            errors.append("weights must sum to 1.0")
    except (KeyError, TypeError, ValueError):
        errors.append("weights must be numeric")

    items = source.get("items")
    if not isinstance(items, list) or len(items) != 11:
        errors.append("exactly 11 reviewed procurement items are required")
        return
    ranks = [item.get("rank") for item in items if isinstance(item, dict)]
    if sorted(ranks) != list(range(1, 12)) or len(set(ranks)) != 11:
        errors.append("procurement ranks must be unique integers 1–11")
    item_ids = [item.get("itemId") for item in items if isinstance(item, dict)]
    if any(not isinstance(value, str) or not value for value in item_ids) or len(set(item_ids)) != 11:
        errors.append("item IDs must be unique non-empty strings")

    outreach = source.get("outreach")
    if not isinstance(outreach, list) or len(outreach) != len(EXPECTED_OUTREACH):
        errors.append("exactly four reviewed outreach records are required")
    else:
        seen_outreach: set[str] = set()
        for record in outreach:
            if not isinstance(record, dict) or set(record) != OUTREACH_KEYS:
                errors.append("outreach record schema differs")
                continue
            item_id = record.get("itemId")
            if item_id in seen_outreach:
                errors.append("outreach item IDs must be unique")
            seen_outreach.add(item_id)
            if item_id not in item_ids or record.get("state") != EXPECTED_OUTREACH.get(item_id):
                errors.append(f"outreach state differs from the approved record for {item_id!r}")
            if not valid_iso_date(record.get("recordedOn")) or record["recordedOn"] > source["asOf"]:
                errors.append(f"outreach date is invalid for {item_id!r}")
            if any(
                not isinstance(record.get(field), str) or not record[field].strip()
                for field in ("noteEn", "noteFi")
            ):
                errors.append(f"bilingual outreach notes are required for {item_id!r}")
            elif (record["noteEn"], record["noteFi"]) != EXPECTED_OUTREACH_NOTES.get(item_id):
                errors.append(f"outreach notes differ from the approved public record for {item_id!r}")
        if seen_outreach != set(EXPECTED_OUTREACH):
            errors.append("outreach item set differs from the approved record")

    for item in items:
        label = f'item {item.get("rank", "?")}' if isinstance(item, dict) else "item"
        if not isinstance(item, dict):
            errors.append(f"{label} must be an object")
            continue
        if item.get("priceType") not in PRICE_TYPES:
            errors.append(f"{label}: invalid priceType")
        if item.get("priceType") == "vendor_quote":
            if item.get("priceAmount") is not None or item.get("currency") is not None:
                errors.append(f"{label}: quote-only item cannot expose a fake numeric price")
            if "quote" not in str(item.get("priceDisplay", "")).casefold():
                errors.append(f"{label}: quote-only price display must say Quote")
        else:
            if not isinstance(item.get("priceAmount"), (int, float)) or item["priceAmount"] <= 0:
                errors.append(f"{label}: public list price must be positive")
            if item.get("currency") not in {"USD", "EUR", "GBP"}:
                errors.append(f"{label}: public list price currency is invalid")
        scores = item.get("scores")
        if not isinstance(scores, dict) or set(scores) != WEIGHT_IDS:
            errors.append(f"{label}: score schema differs")
        elif any(not isinstance(value, (int, float)) or value < 1 or value > 5 for value in scores.values()):
            errors.append(f"{label}: every score must be between 1 and 5")
        else:
            computed = score_item(item, weights)
            if computed < 1 or computed > 5:
                errors.append(f"{label}: weighted score is outside 1–5")
        if not valid_iso_date(item.get("verifiedOn")) or item.get("verifiedOn") > source["asOf"]:
            errors.append(f"{label}: verification date is invalid")
        urls = item.get("sourceUrls")
        if not isinstance(urls, list) or not urls or any(not valid_public_url(url) for url in urls):
            errors.append(f"{label}: official/vendor HTTPS source links are required")
        if item.get("transactionLicenceFit", None) is not None:
            errors.append(f"{label}: transaction licence fit must stay inside the transparent score object")

    if len(source.get("goCriteria", [])) < 5 or len(source.get("stopCriteria", [])) < 5:
        errors.append("at least five go and five stop criteria are required")
    if len(source.get("packageOptions", [])) != 3:
        errors.append("exactly three procurement package options are required")
    combined = "\n".join(text_values(source)).casefold()
    for phrase in FORBIDDEN_PUBLIC_TEXT:
        if phrase in combined:
            errors.append(f"source contains forbidden public text {phrase!r}")
    if EMAIL_PATTERN.search(combined):
        errors.append("source contains an email address")
    if UUID_PATTERN.search(combined):
        errors.append("source contains a UUID-like private reference")
    if "no purchase" not in combined and "mitään ostoa" not in combined:
        errors.append("visible no-purchase boundary is missing")
    if "data-room" not in combined and "datahuone" not in combined:
        errors.append("transaction data-room licence gate is missing")


def validate_outputs(source: dict[str, Any], errors: list[str]) -> None:
    if not OUTPUT_JSON.exists() or OUTPUT_JSON.read_bytes() != render_json(source):
        errors.append("public paid-data JSON is missing or stale")
    elif read_json(OUTPUT_JSON) != normalised(source):
        errors.append("public paid-data JSON differs semantically")
    if not OUTPUT_CSV.exists() or OUTPUT_CSV.read_bytes() != render_csv(source):
        errors.append("public paid-data CSV is missing or stale")
    else:
        rows = list(csv.DictReader(io.StringIO(OUTPUT_CSV.read_text(encoding="utf-8"))))
        if len(rows) != 11 or (rows and list(rows[0]) != CSV_FIELDS):
            errors.append("public paid-data CSV schema or row count differs")
        if any(row["purchaseAuthorised"] != "false" for row in rows):
            errors.append("public paid-data CSV must state purchaseAuthorised=false")
        expected_states = {
            item_id: state for item_id, state in EXPECTED_OUTREACH.items()
        }
        for row, item in zip(rows, sorted(source["items"], key=lambda value: value["rank"])):
            expected_state = expected_states.get(item["itemId"], "not_started")
            if row["outreachState"] != expected_state:
                errors.append(f'public paid-data CSV outreach state differs for {item["itemId"]}')
            if expected_state == "not_started" and any(
                row[field] for field in ("outreachOn", "outreachNoteEn", "outreachNoteFi")
            ):
                errors.append(f'public paid-data CSV exposes unexpected outreach details for {item["itemId"]}')


def validate_workbook(source: dict[str, Any], errors: list[str]) -> None:
    if source.get("version") != CURRENT_DASHBOARD_VERSION:
        errors.append("paid-data dashboard release differs from the reviewed intraday release")
    if source.get("asOf") != WORKBOOK_SNAPSHOT_AS_OF:
        errors.append("retained paid-data workbook is outside the current daily snapshot boundary")
    if not OUTPUT_XLSX.is_file():
        errors.append("reviewed public paid-data XLSX is missing")
        return
    workbook_sha256 = hashlib.sha256(OUTPUT_XLSX.read_bytes()).hexdigest()
    if workbook_sha256 != EXPECTED_XLSX_SHA256:
        errors.append("paid-data XLSX differs from the reviewed release-locked workbook")
    try:
        workbook = load_workbook(OUTPUT_XLSX, read_only=False, data_only=False)
    except Exception as error:  # pragma: no cover - defensive parse boundary
        errors.append(f"cannot parse paid-data XLSX: {error}")
        return
    expected_sheets = [
        "Decision",
        "Priorities",
        "RFP Gate",
        "Avoid",
        "Sources",
        "Response Scorecard",
        "Intake Template",
        "Checks",
    ]
    if workbook.sheetnames != expected_sheets:
        errors.append("paid-data XLSX sheet list differs")
        return
    if any(sheet.sheet_state != "visible" for sheet in workbook.worksheets):
        errors.append("paid-data XLSX cannot contain hidden sheets")
    priority = workbook["Priorities"]
    expected_headers = [
        "Rank",
        "Priority",
        "Vendor",
        "Product",
        "Tier",
        "Price",
        "Decision / Päätös",
        "Annual series",
        "Vape",
        "Countries",
        "Method",
        "Licence",
        "Cost",
        "Weighted score",
        "Alternative group",
        "Source",
    ]
    headers = [priority.cell(5, column).value for column in range(1, 17)]
    if headers != expected_headers:
        errors.append("paid-data XLSX priority headers differ")
    ordered = sorted(source["items"], key=lambda item: item["rank"])
    for index, item in enumerate(ordered, start=6):
        if priority.cell(index, 1).value != item["rank"]:
            errors.append(f"paid-data XLSX rank differs at row {index}")
        if priority.cell(index, 3).value != item["vendor"] or priority.cell(index, 4).value != item["product"]:
            errors.append(f"paid-data XLSX vendor/product differs at row {index}")
        expected_formula = (
            f"=ROUND(H{index}*Decision!$B$17+I{index}*Decision!$B$18+"
            f"J{index}*Decision!$B$19+K{index}*Decision!$B$20+"
            f"L{index}*Decision!$B$21+M{index}*Decision!$B$22,2)"
        )
        if priority.cell(index, 14).value != expected_formula:
            errors.append(f"paid-data XLSX weighted-score formula differs at row {index}")
    euromonitor_item = next(
        item
        for item in source["items"]
        if item["itemId"] == "euromonitor-passport-nicotine"
    )
    expected_euromonitor_decision = "\n".join(
        (
            euromonitor_item["decisionEn"],
            euromonitor_item["conditionsEn"],
            "",
            f'FI: {euromonitor_item["decisionFi"]}',
            euromonitor_item["conditionsFi"],
        )
    )
    if priority["F7"].value != euromonitor_item["priceDisplay"]:
        errors.append("paid-data XLSX Euromonitor quote status differs")
    if priority["G7"].value != expected_euromonitor_decision:
        errors.append("paid-data XLSX Euromonitor decision boundary differs")

    decision = workbook["Decision"]
    if decision["A3"].value != (
        "Independent decision support · No purchase authorised · "
        f"Version {WORKBOOK_SNAPSHOT_VERSION} · Verified {WORKBOOK_SNAPSHOT_AS_OF}"
    ):
        errors.append("paid-data XLSX decision release boundary differs")
    if not isinstance(decision["A10"].value, str) or not all(
        phrase in decision["A10"].value
        for phrase in (
            "Do not activate the conditional Germany extract or buy before explicit purchase authority",
            "Älä aktivoi ehdollista Saksa-otetta tai osta ennen nimenomaista ostovaltuutusta",
            "complete all-in commercial terms",
        )
    ):
        errors.append("paid-data XLSX recommended decision differs")
    recommended_package = next(
        item for item in source["packageOptions"] if item["id"] == "recommended"
    )
    if decision["M17"].value != recommended_package["knownPrice"]:
        errors.append("paid-data XLSX recommended package price status differs")
    if decision["O17"].value != (
        f'{recommended_package["unknownComponentsEn"]}\n\n'
        f'FI: {recommended_package["unknownComponentsFi"]}'
    ):
        errors.append("paid-data XLSX recommended package unknowns differ")

    response = workbook["Response Scorecard"]
    expected_response_headers = [
        "Vendor ID",
        "Vendor",
        "Product",
        "Public outreach state",
        "Sample",
        "Data dictionary / method",
        "Coverage matrix",
        "Official-anchor reconciliation",
        "Transaction-use rights",
        "Total cost / no auto-renew",
        "Annual country series",
        "Metric / scope",
        "Coverage",
        "Methodology",
        "Audit trail",
        "Licence rights",
        "Commercial clarity",
        "Gates passed",
        "Score inputs",
        "Readiness",
        "Weighted score (0–5)",
        "Reviewer note",
        "Source link",
        "Public boundary note",
    ]
    if [response.cell(13, column).value for column in range(1, 25)] != expected_response_headers:
        errors.append("paid-data XLSX response-scorecard headers differ")
    for row, expected_row in enumerate(EXPECTED_RESPONSE_ROWS, start=14):
        reviewed_values = tuple(
            response.cell(row, column).value
            for column in (1, 2, 3, 4, 22, 23, 24)
        )
        if reviewed_values != expected_row:
            errors.append(
                f"paid-data XLSX response identity, public state, reviewer note, "
                f"source or boundary differs at row {row}"
            )
        if any(response.cell(row, column).value != "OPEN" for column in range(5, 11)):
            errors.append(f"paid-data XLSX current vendor gates must remain OPEN at row {row}")
        if any(response.cell(row, column).value is not None for column in range(11, 18)):
            errors.append(f"paid-data XLSX current vendor scores must remain blank at row {row}")
        if response.cell(row, 18).value != f'=COUNTIF(E{row}:J{row},"PASS")&"/6"':
            errors.append(f"paid-data XLSX gate-count formula differs at row {row}")
        if response.cell(row, 19).value != f'=COUNT(K{row}:Q{row})&"/7"':
            errors.append(f"paid-data XLSX score-input formula differs at row {row}")
        readiness_formula = (
            f'=IF(AND(COUNTIF(E{row}:J{row},"PASS")=6,COUNT(K{row}:Q{row})=7),'
            '"READY TO SCORE","NOT SCORED")'
        )
        if response.cell(row, 20).value != readiness_formula:
            errors.append(f"paid-data XLSX response readiness formula differs at row {row}")
        weighted_formula = (
            f'=IF(T{row}="READY TO SCORE",ROUND('
            f"K{row}*$B$25+L{row}*$B$26+M{row}*$B$27+N{row}*$B$28+"
            f"O{row}*$B$29+P{row}*$B$30+Q{row}*$B$31,2),\"\")"
        )
        if response.cell(row, 21).value != weighted_formula:
            errors.append(f"paid-data XLSX response weighted-score formula differs at row {row}")
    expected_response_weights = [0.20, 0.15, 0.15, 0.15, 0.10, 0.15, 0.10]
    if [response.cell(row, 2).value for row in range(25, 32)] != expected_response_weights:
        errors.append("paid-data XLSX response-scorecard weights differ or do not total 100%")

    intake = workbook["Intake Template"]
    expected_intake_headers = [
        "Intake ID",
        "Intake type",
        "Provider / Country",
        "Received on",
        "Evidence type",
        "Years from",
        "Years to",
        "Measure",
        "Supply-chain stage",
        "Product scope",
        "Geography",
        "Currency",
        "Unit",
        "Raw / Derived",
        "Controlled source or file reference",
        "SHA-256",
        "Licence / Terms",
        "Public use allowed",
        "Reviewer",
        "Review status",
        "Confidence",
        "Gaps / Notes",
        "Public citation candidate",
        "Public release decision",
    ]
    if [intake.cell(12, column).value for column in range(1, 25)] != expected_intake_headers:
        errors.append("paid-data XLSX evidence-intake headers differ")
    if any(
        intake.cell(row, column).value is not None
        for row in range(13, 33)
        for column in range(1, 25)
    ):
        errors.append("paid-data XLSX public evidence-intake template must remain blank")

    checks = workbook["Checks"]
    if checks.cell(5, 1).value != (
        '=IF(COUNTIF(F12:F20,"FAIL")=0,"MODEL STATUS: OK / MALLIN TILA: OK",'
        '"MODEL STATUS: REVIEW / MALLIN TILA: TARKISTA")'
    ):
        errors.append("paid-data XLSX overall control formula differs")
    if [checks.cell(row, 3).value for row in range(12, 21)] != [
        1,
        4,
        0,
        0,
        0,
        0,
        0,
        0,
        0,
    ]:
        errors.append("paid-data XLSX visible control expectations differ")
    formula_errors = {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NUM!"}
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.comment is not None:
                    errors.append("paid-data XLSX cannot contain comments")
                if isinstance(cell.value, str):
                    label = f"{sheet.title}!{cell.coordinate}"
                    scan_public_workbook_text(label, cell.value, errors)
                    if cell.value in formula_errors:
                        errors.append(f"paid-data XLSX contains formula error {cell.value}")
                    if cell.value.startswith("=") and "[" in cell.value:
                        errors.append("paid-data XLSX cannot contain external workbook formulas")
                if cell.hyperlink is not None:
                    target = cell.hyperlink.target or cell.hyperlink.location or ""
                    scan_public_workbook_text(
                        f"{sheet.title}!{cell.coordinate} hyperlink",
                        target,
                        errors,
                    )
    workbook.close()

    with zipfile.ZipFile(OUTPUT_XLSX) as package:
        lowered = [name.casefold() for name in package.namelist()]
        forbidden = ("vbaproject", "externallink", "connections", "comments", "embedding", "oleobject")
        for name in lowered:
            if any(part in name for part in forbidden):
                errors.append(f"paid-data XLSX contains forbidden OOXML part {name}")
        for name in package.namelist():
            if not name.casefold().endswith((".xml", ".rels")):
                continue
            try:
                xml_text = package.read(name).decode("utf-8")
            except UnicodeDecodeError:
                errors.append(f"paid-data XLSX OOXML text is not UTF-8: {name}")
                continue
            scan_public_workbook_text(f"OOXML {name}", xml_text, errors)
            if name.casefold().endswith(".rels"):
                try:
                    relationships = ElementTree.fromstring(xml_text)
                except ElementTree.ParseError:
                    errors.append(f"paid-data XLSX relationship XML is invalid: {name}")
                    continue
                for relationship in relationships:
                    target = relationship.attrib.get("Target", "")
                    scan_public_workbook_text(
                        f"OOXML relationship {name}",
                        target,
                        errors,
                    )
                    if relationship.attrib.get("TargetMode") == "External" and not valid_public_url(target):
                        errors.append(
                            f"paid-data XLSX contains a non-public external relationship in {name}"
                        )


def main() -> int:
    errors: list[str] = []
    try:
        source = read_json(SOURCE_PATH)
    except (OSError, json.JSONDecodeError) as error:
        print(f"FAIL: cannot read paid-data source: {error}", file=sys.stderr)
        return 1
    validate_source(source, errors)
    if not errors:
        validate_outputs(source, errors)
        validate_workbook(source, errors)
    if errors:
        for error in errors:
            print(f"FAIL: {error}", file=sys.stderr)
        return 1
    print(
        "PASS: 11-item paid-data shortlist, transparent scores, 3 package options, "
        "go/stop gates, v37 JSON/CSV parity, reviewed v37 daily XLSX snapshot "
        "and no-purchase boundary verified."
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
