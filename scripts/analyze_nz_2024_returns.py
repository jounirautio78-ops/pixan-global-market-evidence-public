#!/usr/bin/env python3
"""Reproduce privacy-safe aggregates from New Zealand's 2024 annual returns.

The raw Ministry workbooks contain respondent-level information and are not
republished by this project. This script validates locally downloaded copies
against the public manifest and emits aggregate results only.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import math
from collections import Counter, defaultdict
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_MANIFEST = REPO_ROOT / "source" / "NZ_2024_WORKBOOK_MANIFEST.json"
DEFAULT_OUTPUT = REPO_ROOT / "source" / "NZ_2024_PRODUCT_SCOPE_AUDIT.json"
MONEY_QUANTUM = Decimal("0.01")
RETURN_CLASSES = ("AIS", "AVP", "Notifier", "RPS")
BUCKETS = (
    "vaping_consumable",
    "vaping_device_or_hardware",
    "vaping_mixed_system",
    "vaping_other_explicit",
    "adjacent_notifiable_product",
    "unresolved_product_type",
)
VAPING_BUCKETS = BUCKETS[:4]
ADJACENT_TERMS = ("smokeless tobacco", "herbal smoking")
CONSUMABLE_TERMS = (
    "vaping substance",
    "e liquid",
    "freebase",
    "nicotine salt",
)
MIXED_SYSTEM_TERMS = ("disposable", "prefilled", "pod", "cartridge")
DEVICE_TERMS = (
    "vaping device",
    "vape device",
    "vapin device",
    "device",
    "kit",
    "tank",
)
GENERIC_VAPING_TERMS = ("vaping", "vape")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--downloads",
        type=Path,
        required=True,
        help="Directory containing the 29 official XLSX workbooks.",
    )
    parser.add_argument(
        "--manifest",
        type=Path,
        default=DEFAULT_MANIFEST,
        help="Public input manifest with file names, sizes and SHA-256 hashes.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help="Aggregate JSON output. No respondent-level values are emitted.",
    )
    return parser.parse_args()


def normalise_header(value: Any) -> str:
    return " ".join(
        str(value or "").strip().casefold().replace("licence", "license").split()
    )


def normalise_product_type(value: Any) -> str:
    return " ".join(str(value or "").casefold().replace("-", " ").split())


def decimal_value(value: Any) -> Decimal | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, float) and not math.isfinite(value):
        return None
    cleaned = str(value).strip().replace(",", "").replace("$", "")
    if not cleaned:
        return None
    try:
        parsed = Decimal(cleaned)
    except InvalidOperation:
        return None
    return parsed if parsed.is_finite() else None


def money(value: Decimal) -> float:
    return float(value.quantize(MONEY_QUANTUM, rounding=ROUND_HALF_UP))


def return_class(file_name: str) -> str:
    for value in RETURN_CLASSES:
        if f"-{value}-" in file_name:
            return value
    raise ValueError(f"Cannot determine return class from {file_name}")


def product_bucket(product_type: Any) -> str:
    value = normalise_product_type(product_type)
    if any(term in value for term in ADJACENT_TERMS):
        return "adjacent_notifiable_product"
    if any(term in value for term in CONSUMABLE_TERMS):
        return "vaping_consumable"
    if any(term in value for term in MIXED_SYSTEM_TERMS):
        return "vaping_mixed_system"
    if any(term in value for term in DEVICE_TERMS):
        return "vaping_device_or_hardware"
    if any(term in value for term in GENERIC_VAPING_TERMS):
        return "vaping_other_explicit"
    return "unresolved_product_type"


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def validate_inputs(downloads: Path, manifest: dict[str, Any]) -> list[dict[str, Any]]:
    expected = manifest.get("files")
    if not isinstance(expected, list) or len(expected) != 29:
        raise ValueError("Manifest must contain exactly 29 file records")
    expected_names = {str(item.get("fileName")) for item in expected}
    actual_names = {path.name for path in downloads.glob("*.xlsx")}
    if actual_names != expected_names:
        missing = sorted(expected_names - actual_names)
        extra = sorted(actual_names - expected_names)
        raise ValueError(f"Workbook set differs from manifest; missing={missing}, extra={extra}")

    checked: list[dict[str, Any]] = []
    for item in sorted(expected, key=lambda record: str(record["fileName"])):
        path = downloads / str(item["fileName"])
        actual_size = path.stat().st_size
        actual_hash = sha256_file(path)
        if actual_size != item.get("bytes"):
            raise ValueError(f"{path.name}: expected {item.get('bytes')} bytes, got {actual_size}")
        if actual_hash != item.get("sha256"):
            raise ValueError(f"{path.name}: SHA-256 differs from the public manifest")
        checked.append(item)
    counts = Counter(return_class(str(item["fileName"])) for item in checked)
    if counts != Counter({"AIS": 1, "AVP": 21, "Notifier": 1, "RPS": 6}):
        raise ValueError(f"Unexpected return-class file counts: {dict(counts)}")
    return checked


def analyse(downloads: Path, manifest_path: Path) -> dict[str, Any]:
    manifest_bytes = manifest_path.read_bytes()
    manifest = json.loads(manifest_bytes)
    checked_inputs = validate_inputs(downloads, manifest)

    rows_by_class: Counter[str] = Counter()
    numeric_sales_by_class: Counter[str] = Counter()
    sales_by_class: defaultdict[str, Decimal] = defaultdict(Decimal)
    rows_by_bucket: Counter[str] = Counter()
    numeric_sales_by_bucket: Counter[str] = Counter()
    sales_by_bucket: defaultdict[str, Decimal] = defaultdict(Decimal)
    deduplicated_sales_by_bucket: defaultdict[str, Decimal] = defaultdict(Decimal)
    repeated_sales_by_bucket: defaultdict[str, Decimal] = defaultdict(Decimal)
    normalised_types_by_bucket: defaultdict[str, set[str]] = defaultdict(set)
    exact_row_counts: Counter[str] = Counter()
    rows_with_numeric_rrp_quantity = 0
    rows_where_sales_equals_rrp_quantity = 0
    rows_where_sales_differs_rrp_quantity = 0
    rows_without_numeric_sales = 0

    for item in checked_inputs:
        workbook_path = downloads / str(item["fileName"])
        kind = return_class(workbook_path.name)
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        for sheet in workbook.worksheets:
            row_iterator = sheet.iter_rows(values_only=True)
            try:
                first_row = next(row_iterator)
            except StopIteration:
                continue
            headers = [normalise_header(value) for value in first_row]
            positions = {header: index for index, header in enumerate(headers) if header}
            product_type_index = positions.get("product type")
            rrp_index = positions.get("rrp")
            quantity_index = positions.get("qty", positions.get("quantity sold"))
            total_index = positions.get("total sales")

            for row in row_iterator:
                if not any(value is not None and str(value).strip() for value in row):
                    continue
                rows_by_class[kind] += 1
                product_type = (
                    row[product_type_index]
                    if product_type_index is not None and product_type_index < len(row)
                    else None
                )
                bucket = product_bucket(product_type)
                rows_by_bucket[bucket] += 1
                normalised_types_by_bucket[bucket].add(
                    normalise_product_type(product_type) or "unspecified"
                )

                total_sales = (
                    decimal_value(row[total_index])
                    if total_index is not None and total_index < len(row)
                    else None
                )
                rrp = (
                    decimal_value(row[rrp_index])
                    if rrp_index is not None and rrp_index < len(row)
                    else None
                )
                quantity = (
                    decimal_value(row[quantity_index])
                    if quantity_index is not None and quantity_index < len(row)
                    else None
                )

                if total_sales is None:
                    rows_without_numeric_sales += 1
                else:
                    numeric_sales_by_class[kind] += 1
                    numeric_sales_by_bucket[bucket] += 1
                    sales_by_class[kind] += total_sales
                    sales_by_bucket[bucket] += total_sales

                if rrp is not None and quantity is not None:
                    rows_with_numeric_rrp_quantity += 1
                    if total_sales is not None:
                        # Preserve the original workbook-QA test exactly. Currency
                        # totals use Decimal; this diagnostic intentionally mirrors
                        # Python/Excel-style binary-number comparison.
                        calculated = float(rrp) * float(quantity)
                        total_for_check = float(total_sales)
                        tolerance = max(0.01, abs(total_for_check) * 1e-9)
                        if abs(total_for_check - calculated) <= tolerance:
                            rows_where_sales_equals_rrp_quantity += 1
                        else:
                            rows_where_sales_differs_rrp_quantity += 1

                row_hash = hashlib.sha256(repr(tuple(row)).encode("utf-8")).hexdigest()
                if total_sales is not None:
                    if exact_row_counts[row_hash]:
                        repeated_sales_by_bucket[bucket] += total_sales
                    else:
                        deduplicated_sales_by_bucket[bucket] += total_sales
                exact_row_counts[row_hash] += 1
        workbook.close()

    total_sales = sum(sales_by_bucket.values(), Decimal())
    vaping_sales = sum((sales_by_bucket[bucket] for bucket in VAPING_BUCKETS), Decimal())
    vaping_deduplicated = sum(
        (deduplicated_sales_by_bucket[bucket] for bucket in VAPING_BUCKETS), Decimal()
    )
    deduplicated_total = sum(deduplicated_sales_by_bucket.values(), Decimal())
    repeated_total = sum(repeated_sales_by_bucket.values(), Decimal())
    repeated_rows = sum(count - 1 for count in exact_row_counts.values() if count > 1)
    rounded_partition = sum(
        (
            value.quantize(MONEY_QUANTUM, rounding=ROUND_HALF_UP)
            for value in sales_by_bucket.values()
        ),
        Decimal(),
    )

    return {
        "schemaVersion": "1.0",
        "reviewDate": "2026-07-26",
        "sourcePage": manifest["sourcePage"],
        "sourceIntegrity": {
            "manifestSha256": hashlib.sha256(manifest_bytes).hexdigest(),
            "filesValidated": len(checked_inputs),
            "downloadedBytes": sum(int(item["bytes"]) for item in checked_inputs),
            "fileCounts": dict(
                sorted(
                    Counter(
                        return_class(str(item["fileName"])) for item in checked_inputs
                    ).items()
                )
            ),
        },
        "publicationBoundary": (
            "Aggregate output only. No respondent, licence code, business identity, "
            "brand, flavour or UPC is emitted."
        ),
        "classificationMethod": {
            "normalisation": "casefold; hyphen to space; collapse whitespace",
            "precedence": [
                "adjacent_notifiable_product",
                "vaping_consumable",
                "vaping_mixed_system",
                "vaping_device_or_hardware",
                "vaping_other_explicit",
                "unresolved_product_type",
            ],
            "terms": {
                "adjacent_notifiable_product": list(ADJACENT_TERMS),
                "vaping_consumable": list(CONSUMABLE_TERMS),
                "vaping_mixed_system": list(MIXED_SYSTEM_TERMS),
                "vaping_device_or_hardware": list(DEVICE_TERMS),
                "vaping_other_explicit": list(GENERIC_VAPING_TERMS),
            },
            "unmatchedTreatment": "unresolved_product_type; excluded from identified vaping",
            "rounding": "Currency aggregates are summed before half-up rounding to two decimals.",
        },
        "rowQuality": {
            "dataRows": sum(rows_by_class.values()),
            "rowsWithNumericTotalSales": sum(numeric_sales_by_class.values()),
            "rowsWithoutNumericTotalSales": rows_without_numeric_sales,
            "rowsWithNumericRrpAndQuantity": rows_with_numeric_rrp_quantity,
            "rowsWhereTotalEqualsRrpTimesQuantity": rows_where_sales_equals_rrp_quantity,
            "rowsWhereTotalDiffersFromRrpTimesQuantity": rows_where_sales_differs_rrp_quantity,
            "exactRepeatedRowSignaturesBeyondFirst": repeated_rows,
        },
        "returnClasses": {
            kind: {
                "files": sum(
                    1 for item in checked_inputs if return_class(str(item["fileName"])) == kind
                ),
                "dataRows": rows_by_class[kind],
                "rowsWithNumericTotalSales": numeric_sales_by_class[kind],
                "reportedTotalSalesNzd": money(sales_by_class[kind]),
            }
            for kind in RETURN_CLASSES
        },
        "productScopeBuckets": {
            bucket: {
                "distinctNormalisedProductTypes": len(normalised_types_by_bucket[bucket]),
                "dataRows": rows_by_bucket[bucket],
                "rowsWithNumericTotalSales": numeric_sales_by_bucket[bucket],
                "reportedTotalSalesNzd": money(sales_by_bucket[bucket]),
                "salesAfterExactRowDeduplicationSensitivityNzd": money(
                    deduplicated_sales_by_bucket[bucket]
                ),
                "salesOnRepeatedExactRowsNzd": money(repeated_sales_by_bucket[bucket]),
            }
            for bucket in BUCKETS
        },
        "publishedAggregates": {
            "allNumericTotalSalesCellsNzd": money(total_sales),
            "identifiedVapingSalesNzd": money(vaping_sales),
            "identifiedVapingDeduplicatedSensitivityNzd": money(vaping_deduplicated),
            "identifiedAdjacentSalesNzd": money(
                sales_by_bucket["adjacent_notifiable_product"]
            ),
            "unresolvedProductTypeSalesNzd": money(
                sales_by_bucket["unresolved_product_type"]
            ),
            "allSalesAfterExactRowDeduplicationSensitivityNzd": money(
                deduplicated_total
            ),
            "salesOnRepeatedExactRowsNzd": money(repeated_total),
            "roundedBucketPartitionMinusRawTotalNzd": money(
                rounded_partition
                - total_sales.quantize(MONEY_QUANTUM, rounding=ROUND_HALF_UP)
            ),
        },
        "interpretationBoundary": {
            "observedValueStage": "AIS and AVP specialist-retailer reported sales only",
            "notifierReportedSalesAdded": False,
            "rpsReportedSalesAdded": False,
            "generalRetailValueStatus": "modelled separately; not part of this observed subtotal",
            "gstBasis": "unknown",
            "nationalCoverage": "incomplete",
            "donorDecision": "not_accepted",
        },
    }


def main() -> None:
    args = parse_args()
    result = analyse(args.downloads.resolve(), args.manifest.resolve())
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        json.dumps(result, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    print(
        json.dumps(
            {
                "files": result["sourceIntegrity"]["filesValidated"],
                "rows": result["rowQuality"]["dataRows"],
                "reportedSalesNzd": result["publishedAggregates"][
                    "allNumericTotalSalesCellsNzd"
                ],
                "identifiedVapingSalesNzd": result["publishedAggregates"][
                    "identifiedVapingSalesNzd"
                ],
                "output": str(args.output),
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
